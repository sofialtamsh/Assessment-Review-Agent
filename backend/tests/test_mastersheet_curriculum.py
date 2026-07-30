"""Curriculum mastersheet (Cheat Sheet + In-class Quiz labels) is parsed into units."""
from __future__ import annotations

import io

from openpyxl import Workbook

from app.ingestion.mastersheet_xlsx import parse_mastersheet_xlsx

HEADER = ["Course", "Topic", "Unit", "Unit Type", "Resources", "Code File",
          "Embedded links", "S3 Links", "s-id", "t-id"]


def _sheet_bytes(rows: list[dict]) -> bytes:
    """Build a mastersheet .xlsx, putting a real hyperlink on the Resources cell
    when a row supplies `res_link` (mirrors how the real sheet links files)."""
    wb = Workbook()
    ws = wb.active
    ws.append(HEADER)
    for r in rows:
        ws.append([r.get("Course", ""), r.get("Topic", ""), r.get("Unit", ""),
                   r.get("Unit Type", ""), r.get("Resources", ""), "",
                   r.get("Embedded links", ""), "", "", ""])
        if r.get("res_link"):
            cell = ws.cell(row=ws.max_row, column=5)  # Resources column
            cell.hyperlink = r["res_link"]
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def test_cheatsheet_and_inclass_quiz_are_captured():
    rows = [
        {"Course": "ML", "Topic": "KNN", "Unit": "KNN", "Unit Type": "Session",
         "Embedded links": "https://docs.google.com/presentation/d/e/ABC/pubembed"},
        {"Course": "ML", "Topic": "KNN", "Unit": "Cheat Sheet | KNN",
         "Unit Type": "Cheat Sheet", "Resources": "KNN - Cheat Sheet",
         "res_link": "https://docs.google.com/document/d/KNNDOC/edit"},
        {"Course": "ML", "Topic": "KNN", "Unit": "In class Quiz | KNN",
         "Unit Type": "In-class Quiz", "Resources": "knn_quiz.xlsx",
         "res_link": "https://drive.google.com/file/d/KNNQUIZ/view"},
        {"Course": "ML", "Topic": "KNN", "Unit": "MCQ Practice | KNN",
         "Unit Type": "MCQ Practice", "Resources": "knn.zip",
         "res_link": "https://drive.google.com/file/d/KNNZIP/view"},
    ]
    units = parse_mastersheet_xlsx(_sheet_bytes(rows))
    assert len(units) == 1
    u = units[0]
    assert u.content_url and "pubembed" in u.content_url          # Session slides
    assert u.tutorial_url and "KNNDOC" in u.tutorial_url          # Cheat Sheet -> reference
    assert u.quiz_doc_url and "KNNQUIZ" in u.quiz_doc_url          # In-class Quiz -> quiz
    assert u.mcq_doc_url and "KNNZIP" in u.mcq_doc_url             # MCQ Practice -> assignment


def test_exam_and_assignment_rows_ignored():
    rows = [
        {"Course": "ML", "Topic": "KNN", "Unit": "Grand Assignment | KNN", "Unit Type": "Exam",
         "Resources": "knn_ga.zip", "res_link": "https://drive.google.com/file/d/GA/view"},
        {"Course": "ML", "Topic": "KNN", "Unit": "Coding Assignment | KNN",
         "Unit Type": "Assignment", "Resources": "code"},
    ]
    # neither is a review source -> no units survive
    assert parse_mastersheet_xlsx(_sheet_bytes(rows)) == []
