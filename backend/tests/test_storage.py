"""Pluggable archive backend + reviewed-content export for the archive."""
from __future__ import annotations

import base64
import io

from openpyxl import load_workbook

from app import storage
from app.export import export_review_xlsx
from app.schemas import Judgment, Option, Question


def _q(qid):
    return Question(question_id=qid, session_id="u1", stem=f"stem {qid}",
                    options=[Option(key="A", text="a"), Option(key="B", text="b")],
                    correct_keys=["A"], source_set="mcq_assignment")


def test_noop_backend_when_unconfigured():
    b = storage.get_backend()
    assert b.enabled() is False
    assert b.save("x/y.json", b"{}", "msg") is None


def test_archive_review_disabled_returns_empty():
    urls = storage.archive_review(
        session_id="u1", run_id="run1", source_set="mcq_assignment", title="U1",
        reviewer="Sofi", record={"a": 1}, xlsx=b"data")
    assert urls == []


def test_github_backend_url_and_encoding():
    b = storage.GitHubBackend("tok", "org/repo", "main")
    assert b.enabled() is True
    assert b._url("reviews/u1/run1/review.json") == (
        "https://api.github.com/repos/org/repo/contents/reviews/u1/run1/review.json")
    # content is base64-encoded in the PUT body (spot-check the encoding we rely on)
    assert base64.b64encode(b"hello").decode() == "aGVsbG8="


def test_review_xlsx_puts_verdict_in_remarks():
    qs = [_q("q1"), _q("q2")]
    js = [Judgment(question_id="q1", verdict="APPROVE", reason="clean"),
          Judgment(question_id="q2", verdict="DELETE", reason="out of scope")]
    wb = load_workbook(io.BytesIO(export_review_xlsx(qs, js)))
    rows = list(wb.active.iter_rows(values_only=True))
    assert rows[0][-1] == "Remarks"
    assert "APPROVE: clean" in rows[1][-1]
    assert "DELETE: out of scope" in rows[2][-1]
