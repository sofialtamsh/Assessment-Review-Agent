"""Auth (shared password), review-history guardrail, activity feed, and export format."""
from __future__ import annotations

import io

from openpyxl import load_workbook

from app import auth, store
from app.export import export_cleaned_xlsx
from app.schemas import Option, Question


# --- auth ------------------------------------------------------------------ #
def test_login_success_and_token():
    r = auth.login("Sofi", "admin@123")
    assert r["name"] == "Sofi"
    assert auth.verify_token("Sofi", r["token"])
    assert not auth.verify_token("Sofi", "bogus")


def test_login_wrong_password():
    import pytest
    with pytest.raises(ValueError):
        auth.login("Sofi", "nope")


# --- review history / guardrail ------------------------------------------- #
def _q(qid, stem):
    return Question(question_id=qid, session_id="u1", stem=stem,
                    options=[Option(key="A", text="a"), Option(key="B", text="b")],
                    correct_keys=["A"], source_set="mcq_assignment")


def test_summary_and_prior_by_unit_and_content():
    qs = [_q("q1", "what is x"), _q("q2", "what is y")]
    report = {"total_questions": 2, "pass_rate": 0.5,
              "verdict_counts": {"APPROVE": 1, "REVISE": 1}, "rubric_applied": False}
    store.save_review_summary(run_id="run_a", session_id="u1", source_set="mcq_assignment",
                              title="Unit 1", reviewer="Ravi", report=report, questions=qs)

    # same unit + set -> found
    prior = store.find_prior_reviews("u1", "mcq_assignment")
    assert prior and prior[0]["reviewer"] == "Ravi"
    assert prior[0]["verdict_counts"]["APPROVE"] == 1

    # identical question content under a DIFFERENT session id -> still found via hash
    chash = store.question_set_hash(qs)
    prior2 = store.find_prior_reviews("other_unit", "mcq_assignment", content_hash=chash)
    assert prior2 and prior2[0]["run_id"] == "run_a"

    # a genuinely different unit + different content -> nothing
    assert store.find_prior_reviews("u_zzz", "in_class_quiz") == []


def test_question_set_hash_order_independent():
    a = [_q("q1", "alpha"), _q("q2", "beta")]
    b = [_q("q2", "beta"), _q("q1", "alpha")]
    assert store.question_set_hash(a) == store.question_set_hash(b)


def test_activity_lists_runs_with_reviewer():
    from app.jobs import manager
    rid = manager.create_run("u_activity", "mcq_assignment", reviewer="Meena")
    act = store.list_activity(50)
    mine = [a for a in act if a["run_id"] == rid]
    assert mine and mine[0]["reviewer"] == "Meena"


def test_clear_units_replaces_but_keeps_evals_and_others():
    from app.models import SessionRow
    from app.db import get_session as _gs
    from app.schemas import UnitSpec

    store.save_units([UnitSpec(unit_id="clr_u1", unit="Old A"),
                      UnitSpec(unit_id="clr_u2", unit="Old B")], owner="Ravi")
    # an eval session + another user's unit must survive Ravi's re-ingest
    with _gs() as db:
        db.add(SessionRow(session_id="eval_keepme", owner="Ravi", unit="Eval"))
        db.commit()
    store.save_units([UnitSpec(unit_id="clr_other", unit="Other")], owner="Zoe")

    store.clear_units("Ravi")                       # fresh ingest for Ravi
    store.save_units([UnitSpec(unit_id="clr_new", unit="New")], owner="Ravi")

    ravi = {u["unit_id"] for u in store.list_units("Ravi")}
    assert ravi == {"clr_new"}                       # only the new ingest remains
    with _gs() as db:
        assert db.get(SessionRow, "eval_keepme") is not None   # eval preserved
    assert {u["unit_id"] for u in store.list_units("Zoe")} == {"clr_other"}  # other user safe


def test_insights_aggregates_reviews_and_issues():
    from app.db import get_session as _gs
    from app.models import FindingRow

    qs = [_q("iq1", "insight stem")]
    store.save_review_summary(run_id="ins_run", session_id="ins_u", source_set="mcq_assignment",
                              title="Ins", reviewer="Nadia",
                              report={"total_questions": 10, "pass_rate": 0.8,
                                      "verdict_counts": {"APPROVE": 8, "REVISE": 2}},
                              questions=qs)
    with _gs() as db:
        db.add(FindingRow(run_id="ins_run", question_id="iq1", phase="phase3_ambiguity",
                          check_name="option_ambiguity", verdict="FAIL"))
        db.add(FindingRow(run_id="ins_run", question_id="iq1", phase="phase4_scope",
                          check_name="out_of_scope", verdict="WARN"))
        db.commit()

    ins = store.insights()
    assert ins["total_reviews"] >= 1
    assert ins["total_questions"] >= 10
    assert 0 <= ins["avg_approval_pct"] <= 100
    checks = {i["check"] for i in ins["top_issues"]}
    assert "option_ambiguity" in checks and "out_of_scope" in checks
    assert any(r["reviewer"] == "Nadia" for r in ins["by_reviewer"])


def test_units_scoped_per_owner_but_activity_shared():
    from app.jobs import manager
    from app.schemas import UnitSpec

    store.save_units([UnitSpec(unit_id="iso_u1", unit="U1")], owner="Alice")
    store.save_units([UnitSpec(unit_id="iso_u2", unit="U2")], owner="Bob")
    alice = {u["unit_id"] for u in store.list_units("Alice")}
    bob = {u["unit_id"] for u in store.list_units("Bob")}
    assert "iso_u1" in alice and "iso_u2" not in alice   # Alice sees only hers
    assert "iso_u2" in bob and "iso_u1" not in bob        # Bob sees only his
    # reviews / activity stay GLOBAL — everyone's reviewed data is visible
    manager.create_run("iso_u1", "mcq_assignment", reviewer="Alice")
    manager.create_run("iso_u2", "mcq_assignment", reviewer="Bob")
    reviewers = {a["reviewer"] for a in store.list_activity(50)}
    assert {"Alice", "Bob"} <= reviewers


# --- export format --------------------------------------------------------- #
def test_reviewed_export_format():
    qs = [Question(question_id="q1", session_id="u1", stem="Pick the operator",
                   options=[Option(key="A", text="/"), Option(key="B", text="//"),
                            Option(key="C", text="%"), Option(key="D", text="**")],
                   correct_keys=["B"], explanation="floor div", difficulty="easy",
                   subtopics=["arithmetic"], source_set="examination")]
    wb = load_workbook(io.BytesIO(export_cleaned_xlsx(qs)))
    ws = wb.active
    assert ws.title == "MCQs"
    rows = list(ws.iter_rows(values_only=True))
    assert list(rows[0]) == ["S. No", "question content", "Option A", "Option B",
                             "Option C", "Option D", "Explanation", "Key", "SUB TOPIC",
                             "Difficulty", "pool", "Image (if any)", "Remarks"]
    r = rows[1]
    assert r[0] == 1 and r[1] == "Pick the operator"
    assert r[7] == "Option B"          # correct key rendered as "Option B"
    assert r[8] == "arithmetic" and r[9] == "easy"
