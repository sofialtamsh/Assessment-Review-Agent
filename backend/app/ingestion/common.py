"""Shared tabular-file helpers (CSV / XLSX) with no heavy pandas dependency."""
from __future__ import annotations

import csv
import io
from typing import Any


def _norm_key(k: str) -> str:
    return str(k or "").strip().lower().replace(" ", "_").replace("-", "_")


def read_all_tables(data: bytes, filename: str) -> list[tuple[str, list[dict[str, Any]]]]:
    """Return [(sheet_name, [row dicts])] for EVERY sheet, with normalized header keys.

    A .csv is a single unnamed sheet; a .xlsx/.xlsm yields one entry per worksheet
    (so a multi-tab question pool — MCQs, Fill-in-the-blanks, … — is read whole).
    """
    name = (filename or "").lower()
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        sheets = _read_xlsx_all(data)
    else:
        sheets = [("", _read_csv(data))]
    return [
        (sheet, [{_norm_key(k): ("" if v is None else v) for k, v in row.items()} for row in rows])
        for sheet, rows in sheets
    ]


def read_table(data: bytes, filename: str) -> list[dict[str, Any]]:
    """Return the FIRST sheet's rows with normalized (snake_case) header keys.

    Supports .csv and .xlsx. Falls back to CSV parsing for unknown extensions.
    """
    tables = read_all_tables(data, filename)
    return tables[0][1] if tables else []


def _read_csv(data: bytes) -> list[dict[str, Any]]:
    text = data.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    return [dict(r) for r in reader]


def _rows_from_sheet(ws) -> list[dict[str, Any]]:
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return []
    header = [str(h) if h is not None else f"col{i}" for i, h in enumerate(header)]
    out: list[dict[str, Any]] = []
    for row in rows_iter:
        if row is None or all(c is None for c in row):
            continue
        out.append({header[i]: row[i] if i < len(row) else None for i in range(len(header))})
    return out


def _read_xlsx_all(data: bytes) -> list[tuple[str, list[dict[str, Any]]]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    out: list[tuple[str, list[dict[str, Any]]]] = []
    for ws in wb.worksheets:
        rows = _rows_from_sheet(ws)
        if rows:
            out.append((ws.title, rows))
    return out


def split_list(value: Any) -> list[str]:
    """Split a cell that may hold a delimited list (comma/semicolon/pipe)."""
    if value is None:
        return []
    if isinstance(value, list):
        return [str(v).strip() for v in value if str(v).strip()]
    s = str(value).strip()
    if not s:
        return []
    # newline-separated lists (e.g. a "What to Cover" cell) split first
    if "\n" in s:
        return [p.strip(" -•\t") for p in s.splitlines() if p.strip(" -•\t")]
    for sep in [";", "|", ","]:
        if sep in s:
            return [p.strip() for p in s.split(sep) if p.strip()]
    return [s]


def first(row: dict[str, Any], *keys: str, default: Any = "") -> Any:
    for k in keys:
        if k in row and str(row[k]).strip() != "":
            return row[k]
    return default
