"""Tutorial cheat-sheet ingestion.

A unit's Tutorial is a Google Sheet (exported as .xlsx) whose `TutorialStep` sheet
holds the teaching content: a `content` column with the tutorial written in Markdown
(one big cell per step, ordered by `order`). We extract that text so it can be
chunked and used as REFERENCE MATERIAL for scope-checking, ALONGSIDE the slides —
never as a question source.
"""
from __future__ import annotations

import io
import re

from ..schemas import Chunk
from .content import chunk_segments

_TAG_RE = re.compile(r"</?[A-Za-z][^>]*>")            # <IndexList>, <IndexItem ...>, ...
_MD_IMG = re.compile(r"!\[[^\]]*\]\([^)]*\)")           # markdown images


def _clean(text: str) -> str:
    """Light cleanup: drop custom/HTML tags and image markup, keep the prose + code."""
    text = _MD_IMG.sub(" ", text)
    text = _TAG_RE.sub(" ", text)
    return text.strip()


def extract_tutorial_segments(data: bytes) -> list[tuple[str, str]]:
    """Return [(source_ref, text)] from the TutorialStep sheet's `content` column.

    Falls back to any sheet that has a `content` column if `TutorialStep` is absent.
    """
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)

    ws = None
    for name in wb.sheetnames:
        if name.strip().lower() == "tutorialstep":
            ws = wb[name]
            break
    if ws is None:  # fallback: first sheet that has a "content" header
        for name in wb.sheetnames:
            cand = wb[name]
            header = _header(cand)
            if "content" in header:
                ws = cand
                break
    if ws is None:
        return []

    header = _header(ws)
    c_content = header.get("content")
    c_order = header.get("order")
    if c_content is None:
        return []

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None:
            continue
        raw = row[c_content] if c_content < len(row) else None
        text = _clean(str(raw)) if raw is not None else ""
        if not text:
            continue
        order = row[c_order] if (c_order is not None and c_order < len(row)) else None
        try:
            order_val = int(order)
        except (TypeError, ValueError):
            order_val = len(rows) + 1
        rows.append((order_val, text))

    rows.sort(key=lambda t: t[0])
    if len(rows) == 1:
        return [("Tutorial", rows[0][1])]
    return [(f"Tutorial step {i}", text) for i, (_, text) in enumerate(rows, start=1)]


def _header(ws) -> dict[str, int]:
    """Map normalized header name -> column index from the first row."""
    try:
        first = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    except StopIteration:
        return {}
    out: dict[str, int] = {}
    for i, h in enumerate(first or []):
        if h is None:
            continue
        out[str(h).strip().lower().replace(" ", "_")] = i
    return out


def parse_tutorial(session_id: str, data: bytes) -> list[Chunk]:
    """Full pipeline: xlsx bytes -> cleaned segments -> chunks (keyed to the session)."""
    return chunk_segments(session_id, extract_tutorial_segments(data))
