"""Parse an XLSX mastersheet, extracting the cell HYPERLINKS (which CSV export
drops), and aggregate rows into logical units.

Each topic (e.g. "Naive Bayes Classifier") appears across several rows, one per
Unit Type. Two layouts are supported and grouped the same way:

  * Session      -> slide content   (Unit = "<topic>")
  * Tutorial     -> tutorial cheat-sheet, used as EXTRA reference content
                    (Unit = "Tutorial | <topic>")
  * MCQ Practice -> the MCQ assignment (a Drive .zip or a Doc)
                    (Unit = "MCQ Practice | <topic>")

We strip the "<type> | " prefix so all three rows collapse into one UnitSpec that
carries the links to fetch content + questions. Links are read from whichever
column holds a hyperlink (Resources / Embedded links / PPT / S3), so both the
newer "Resources" sheets and older "PPT"/"Embedded links" sheets work.
Coding Assignment / Exam / Project rows are ignored for review sourcing.
"""
from __future__ import annotations

import io
import re

from ..schemas import UnitSpec
from .common import split_list

_SLUG = re.compile(r"[^a-z0-9]+")
_DOC_LINK = re.compile(r"docs\.google\.com/document/", re.I)  # Google Doc URL
# unit-type prefixes we strip from "<type> | <topic>" to get the grouping key
_TYPE_PREFIXES = {
    "session", "tutorial", "mcq practice", "mcq", "coding practice", "assignment",
    "grand assignment", "exam", "project", "reading material", "coding assignment",
}


def parse_mastersheet_xlsx(data: bytes) -> list[UnitSpec]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), data_only=True)
    ws = wb.active
    header = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    idx = {h.lower(): i for i, h in enumerate(header)}

    def col(*names: str) -> int | None:
        for n in names:
            if n.lower() in idx:
                return idx[n.lower()]
        return None

    c_course = col("course")
    c_topic = col("topic")
    c_unit = col("unit")
    c_cover = col("what to cover", "subtopics")
    c_type = col("unit type", "type")
    c_ppt = col("ppt")
    c_embed = col("embedded links", "embedded link")
    c_resources = col("resources")
    c_s3 = col("s3 links", "s3 link")
    c_sid = col("s-id", "s_id", "sid")

    units: dict[str, UnitSpec] = {}
    for row in ws.iter_rows(min_row=2):
        unit_name = _cell(row, c_unit)
        if not unit_name:
            continue
        utype = _cell(row, c_type).lower()
        topic_key = _topic_key(unit_name)
        course = _cell(row, c_course)
        key = _SLUG.sub("-", f"{course}-{topic_key}".lower()).strip("-")
        u = units.setdefault(key, UnitSpec(
            unit_id=key, course=course, module=_cell(row, c_topic), unit=topic_key,
        ))
        if not u.module:
            u.module = _cell(row, c_topic)

        # the row's link: for slides prefer the public "Embedded links"/pubembed,
        # otherwise the first hyperlink found in Resources / PPT / S3.
        embed_link = _hyperlink(row, c_embed) or _cell_url(row, c_embed)
        res_link = _hyperlink(row, c_resources)
        ppt_link = _hyperlink(row, c_ppt)
        s3_link = _hyperlink(row, c_s3) or _cell_url(row, c_s3)
        any_link = res_link or ppt_link or embed_link or s3_link

        if "session" in utype:
            cover = _cell(row, c_cover)
            if cover:
                u.subtopics = split_list(cover)
            # slides: published/embedded link is publicly fetchable; fall back to any
            u.content_url = embed_link or ppt_link or s3_link or res_link or u.content_url
            u.s_id = _cell(row, c_sid) or u.s_id
            u.unit = unit_name  # bare topic name reads best as the display name
        elif "mcq" in utype:            # "MCQ Practice" -> the assignment (zip or doc)
            u.mcq_doc_url = u.mcq_doc_url or any_link
        elif "tutorial" in utype:       # tutorial: reference content (+ in-class MCQs if a doc)
            u.tutorial_url = u.tutorial_url or any_link
            # A Tutorial *doc* also carries the in-class MCQs at the end, so it
            # doubles as the in-class quiz source. A Tutorial *sheet*/.xlsx is
            # reference content only (no MCQs to parse out of it).
            if any_link and _looks_like_doc(any_link):
                u.quiz_doc_url = u.quiz_doc_url or any_link
        # (Coding Assignment / Exam / Project rows are ignored for review sourcing)

    # keep only units that have at least a question source or some content
    return [u for u in units.values()
            if u.mcq_doc_url or u.quiz_doc_url or u.content_url or u.tutorial_url]


def _looks_like_doc(url: str) -> bool:
    """True if the link is a Google Doc or a direct .doc/.docx — the tutorial doc
    whose tail holds the in-class MCQs (a .xlsx/Sheet tutorial has no MCQs)."""
    u = (url or "").strip().lower()
    return bool(_DOC_LINK.search(u)) or u.split("?")[0].endswith((".doc", ".docx"))


def _topic_key(unit_name: str) -> str:
    """Strip a leading "<unit type> | " prefix so all rows of a topic group together."""
    if "|" in unit_name:
        left, right = unit_name.split("|", 1)
        if left.strip().lower() in _TYPE_PREFIXES:
            return right.strip()
    return unit_name.strip()


def _cell(row, i: int | None) -> str:
    if i is None or i >= len(row):
        return ""
    v = row[i].value
    return str(v).strip() if v is not None else ""


def _cell_url(row, i: int | None) -> str | None:
    """Plain cell value, only if it looks like a URL."""
    v = _cell(row, i)
    return v if v.lower().startswith("http") else None


def _hyperlink(row, i: int | None) -> str | None:
    if i is None or i >= len(row):
        return None
    c = row[i]
    if c.hyperlink and c.hyperlink.target:
        return c.hyperlink.target
    # sometimes the URL is the plain value
    v = str(c.value or "").strip()
    return v if v.lower().startswith("http") else None
